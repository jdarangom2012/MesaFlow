from datetime import datetime, timedelta
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg, Count, DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce, ExtractHour
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from accounts.permissions import ADMIN, SUPERADMIN, role_required, tenant_filter
from orders.models import Order, OrderItem
from tables.models import RestaurantTable


ZERO_MONEY = Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))
ORDER_STATUS_FILTERS = ['ALL'] + [status for status, _label in Order.STATUS_CHOICES]
PAYMENT_METHOD_FILTERS = ['ALL'] + [method for method, _label in Order.PAYMENT_METHOD_CHOICES]


def _parse_date(value):
    if not value:
        return None

    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def get_period_bounds(period):
    today = timezone.localdate()

    if period == 'week':
        start_date = today - timedelta(days=today.weekday())
    elif period == 'month':
        start_date = today.replace(day=1)
    else:
        period = 'today'
        start_date = today

    return period, start_date, today


def _get_filters(request):
    period = request.GET.get('period', 'today')
    period, period_start, today = get_period_bounds(period)
    start_date = _parse_date(request.GET.get('start_date')) or period_start
    end_date = _parse_date(request.GET.get('end_date')) or today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    selected_status = request.GET.get('status', 'ALL').upper()
    selected_payment = request.GET.get('payment_method', 'ALL').upper()
    selected_table = request.GET.get('table', 'ALL')

    if selected_status not in ORDER_STATUS_FILTERS:
        selected_status = 'ALL'

    if selected_payment not in PAYMENT_METHOD_FILTERS:
        selected_payment = 'ALL'

    return {
        'period': period,
        'today': today,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': selected_status,
        'selected_payment': selected_payment,
        'selected_table': selected_table,
    }


def _filtered_orders(request):
    filters = _get_filters(request)
    start_at = timezone.make_aware(datetime.combine(filters['start_date'], datetime.min.time()))
    end_at = timezone.make_aware(datetime.combine(filters['end_date'], datetime.max.time()))
    orders = (
        Order.objects
        .filter(created_at__gte=start_at, created_at__lte=end_at, **tenant_filter(request))
        .select_related('table', 'restaurant')
        .prefetch_related('items__product')
        .order_by('-created_at')
    )

    if filters['selected_status'] != 'ALL':
        orders = orders.filter(status=filters['selected_status'])

    if filters['selected_payment'] != 'ALL':
        orders = orders.filter(payment_method=filters['selected_payment'])

    if filters['selected_table'] != 'ALL':
        orders = orders.filter(table_id=filters['selected_table'])

    return orders, filters


def _build_reports_context(request):
    orders, filters = _filtered_orders(request)
    paid_orders = orders.filter(status='PAID')
    tables = RestaurantTable.objects.filter(**tenant_filter(request)).order_by('name')
    payment_method_labels = dict(Order.PAYMENT_METHOD_CHOICES)
    status_labels = dict(Order.STATUS_CHOICES)

    kpis = orders.aggregate(
        total_orders=Count('id'),
        paid_orders=Count('id', filter=Q(status='PAID')),
        total_sales=Coalesce(Sum('total', filter=Q(status='PAID')), ZERO_MONEY),
        total_tax=Coalesce(Sum('tax', filter=Q(status='PAID')), ZERO_MONEY),
        average_ticket=Coalesce(Avg('total', filter=Q(status='PAID')), ZERO_MONEY),
    )

    top_products = (
        OrderItem.objects
        .filter(order__in=paid_orders)
        .values('product__name')
        .annotate(
            quantity_sold=Coalesce(Sum('quantity'), Value(0)),
            revenue=Coalesce(Sum('subtotal'), ZERO_MONEY),
        )
        .order_by('-quantity_sold', '-revenue')[:8]
    )

    payment_methods = list(
        paid_orders
        .values('payment_method')
        .annotate(total=Coalesce(Sum('total'), ZERO_MONEY), orders=Count('id'))
        .order_by('-total')
    )
    total_sales = kpis['total_sales']
    for method in payment_methods:
        method['label'] = payment_method_labels.get(method['payment_method'], 'Sin metodo')
        method['percentage'] = float((method['total'] / total_sales) * 100) if total_sales else 0

    sales_by_hour = (
        paid_orders
        .annotate(hour=ExtractHour('paid_at'))
        .values('hour')
        .annotate(total=Coalesce(Sum('total'), ZERO_MONEY), orders=Count('id'))
        .order_by('hour')
    )
    hour_totals = {row['hour']: float(row['total']) for row in sales_by_hour if row['hour'] is not None}
    hour_labels = [f'{hour:02d}:00' for hour in range(24)]
    hour_values = [hour_totals.get(hour, 0) for hour in range(24)]

    tables_usage = (
        paid_orders
        .values('table__name')
        .annotate(orders=Count('id'), total=Coalesce(Sum('total'), ZERO_MONEY))
        .order_by('-orders', '-total')[:8]
    )

    paginator = Paginator(orders, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    query_params = request.GET.copy()
    query_params.pop('page', None)
    export_query = query_params.urlencode()

    chart_data = {
        'salesByHour': {
            'labels': hour_labels,
            'values': hour_values,
        },
        'paymentMethods': {
            'labels': [row['label'] for row in payment_methods],
            'values': [float(row['total']) for row in payment_methods],
        },
        'topProducts': {
            'labels': [row['product__name'] for row in top_products],
            'values': [int(row['quantity_sold']) for row in top_products],
        },
        'tablesUsage': {
            'labels': [row['table__name'] for row in tables_usage],
            'values': [int(row['orders']) for row in tables_usage],
        },
    }

    return {
        'restaurant': request.restaurant,
        'period': filters['period'],
        'today': filters['today'],
        'start_date': filters['start_date'],
        'end_date': filters['end_date'],
        'selected_status': filters['selected_status'],
        'selected_payment': filters['selected_payment'],
        'selected_table': filters['selected_table'],
        'status_filters': ORDER_STATUS_FILTERS,
        'status_labels': status_labels,
        'payment_filters': PAYMENT_METHOD_FILTERS,
        'payment_method_labels': payment_method_labels,
        'tables': tables,
        'orders': orders,
        'page_obj': page_obj,
        'export_query': export_query,
        'total_sales': kpis['total_sales'],
        'total_orders': kpis['total_orders'],
        'paid_orders': kpis['paid_orders'],
        'average_ticket': kpis['average_ticket'],
        'total_tax': kpis['total_tax'],
        'tips_total': 0,
        'top_products': top_products,
        'payment_methods': payment_methods,
        'tables_usage': tables_usage,
        'chart_data': chart_data,
    }


@login_required
@role_required(SUPERADMIN, ADMIN)
def reports_dashboard(request):
    return render(request, 'reports/index.html', _build_reports_context(request))


@login_required
@role_required(SUPERADMIN, ADMIN)
def export_reports_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    context = _build_reports_context(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    header_fill = PatternFill('solid', fgColor='00D4FF')
    header_font = Font(color='07111F', bold=True)

    summary_rows = [
        ('Restaurante', context['restaurant'].name if context['restaurant'] else 'MesaFlow'),
        ('Fecha inicio', context['start_date'].isoformat()),
        ('Fecha fin', context['end_date'].isoformat()),
        ('Ventas', float(context['total_sales'])),
        ('Ordenes', context['total_orders']),
        ('Ticket promedio', float(context['average_ticket'])),
        ('IVA total', float(context['total_tax'])),
        ('Propinas', float(context['tips_total'])),
    ]
    for row in summary_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    orders_ws = wb.create_sheet('Ordenes')
    orders_headers = ['Orden', 'Mesa', 'Estado', 'Fecha', 'Metodo pago', 'Subtotal', 'IVA', 'Total']
    orders_ws.append(orders_headers)
    for cell in orders_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for order in context['orders']:
        orders_ws.append([
            order.id,
            order.table.name,
            order.get_status_display(),
            timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M'),
            order.get_payment_method_display() if order.payment_method else 'Sin metodo',
            float(order.subtotal),
            float(order.tax),
            float(order.total),
        ])

    products_ws = wb.create_sheet('Productos')
    products_ws.append(['Producto', 'Cantidad', 'Ingresos'])
    for cell in products_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for product in context['top_products']:
        products_ws.append([product['product__name'], product['quantity_sold'], float(product['revenue'])])

    payments_ws = wb.create_sheet('Metodos pago')
    payments_ws.append(['Metodo', 'Ordenes', 'Total', 'Porcentaje'])
    for cell in payments_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for method in context['payment_methods']:
        payments_ws.append([method['label'], method['orders'], float(method['total']), round(method['percentage'], 2)])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'mesaflow-reporte-{context["start_date"]}-{context["end_date"]}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
@role_required(SUPERADMIN, ADMIN)
def export_reports_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    context = _build_reports_context(request)
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('MesaFlow Reporte Ejecutivo', styles['Title']),
        Paragraph(context['restaurant'].name if context['restaurant'] else 'MesaFlow', styles['Heading2']),
        Paragraph(f'{context["start_date"]} a {context["end_date"]}', styles['Normal']),
        Spacer(1, 16),
    ]

    summary_data = [
        ['Ventas', 'Ordenes', 'Ticket promedio', 'IVA', 'Propinas'],
        [
            f'${context["total_sales"]:,.0f}',
            str(context['total_orders']),
            f'${context["average_ticket"]:,.0f}',
            f'${context["total_tax"]:,.0f}',
            f'${context["tips_total"]:,.0f}',
        ],
    ]
    summary_table = Table(summary_data, colWidths=[92, 82, 112, 92, 92])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00D4FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#07111F')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F4FAFF')),
        ('GRID', (0, 0), (-1, -1), .5, colors.HexColor('#B9C7D6')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.extend([summary_table, Spacer(1, 18), Paragraph('Ordenes', styles['Heading2'])])

    order_rows = [['Orden', 'Mesa', 'Estado', 'Fecha', 'Metodo', 'Total']]
    for order in context['orders'][:30]:
        order_rows.append([
            f'#{order.id}',
            order.table.name,
            order.get_status_display(),
            timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M'),
            order.get_payment_method_display() if order.payment_method else 'Sin metodo',
            f'${order.total:,.0f}',
        ])
    if len(order_rows) == 1:
        order_rows.append(['Sin datos', '', '', '', '', ''])

    orders_table = Table(order_rows, colWidths=[54, 70, 78, 105, 82, 80], repeatRows=1)
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#081028')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), .35, colors.HexColor('#B9C7D6')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.extend([orders_table, Spacer(1, 18), Paragraph('Top productos', styles['Heading2'])])

    product_rows = [['Producto', 'Cantidad', 'Ingresos']]
    for product in context['top_products']:
        product_rows.append([product['product__name'], product['quantity_sold'], f'${product["revenue"]:,.0f}'])
    if len(product_rows) == 1:
        product_rows.append(['Sin datos', '', ''])

    product_table = Table(product_rows, colWidths=[250, 80, 120], repeatRows=1)
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#081028')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), .35, colors.HexColor('#B9C7D6')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(product_table)
    doc.build(story)
    output.seek(0)

    filename = f'mesaflow-reporte-{context["start_date"]}-{context["end_date"]}.pdf'
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
