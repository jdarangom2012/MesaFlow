from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from accounts.permissions import ADMIN, CAJERO, SUPERADMIN, role_required, tenant_filter, user_is_superadmin
from restaurants.models import Restaurant

from .models import CashMovement, CashRegisterSession
from .services import get_open_session, get_session_totals


def parse_money(value):
    try:
        amount = Decimal(value or '0')
    except (InvalidOperation, TypeError):
        amount = Decimal('0')

    return max(amount, Decimal('0'))


def _get_session_for_export(request, session_id):
    return get_object_or_404(
        CashRegisterSession.objects.select_related('restaurant', 'opened_by', 'closed_by'),
        id=session_id,
        **tenant_filter(request),
    )


def _session_context(session):
    totals = get_session_totals(session)
    movements = (
        CashMovement.objects
        .filter(restaurant=session.restaurant, session=session)
        .select_related('order', 'user')
        .order_by('-created_at')
    )

    if session.status == CashRegisterSession.STATUS_CLOSED:
        totals['expected_total'] = session.expected_total
        totals['expected_cash'] = session.expected_cash or session.expected_total
        totals['cash_total'] = session.total_cash
        totals['card_total'] = session.total_card
        totals['digital_total'] = session.total_qr
        totals['qr_total'] = session.total_qr
        totals['sales_total'] = session.total_sales
        totals['tips_total'] = session.total_tips
        totals['tax_total'] = session.total_tax
        totals['expense_total'] = session.total_expenses

    return totals, movements


def _resolve_cash_restaurant(request):
    if getattr(request, 'restaurant', None):
        return request.restaurant, Restaurant.objects.none()

    if not user_is_superadmin(request.user):
        return None, Restaurant.objects.none()

    restaurants = Restaurant.objects.filter(is_active=True).order_by('name')
    restaurant_id = request.POST.get('restaurant_id') or request.GET.get('restaurant_id')

    if restaurant_id:
        restaurant = restaurants.filter(id=restaurant_id).first()
    else:
        restaurant = restaurants.first()

    return restaurant, restaurants


def _dashboard_redirect(restaurant=None):
    url = reverse('cash_register:dashboard')

    if restaurant:
        return redirect(f'{url}?restaurant_id={restaurant.id}')

    return redirect(url)


@login_required
@role_required(SUPERADMIN, ADMIN, CAJERO)
@require_http_methods(['GET', 'POST'])
def cash_register_dashboard(request):
    restaurant, restaurants = _resolve_cash_restaurant(request)

    if not restaurant:
        return render(request, 'auth/403.html', {
            'message': 'Selecciona o asigna un restaurante para operar caja.',
        }, status=403)

    if request.method == 'POST':
        action = request.POST.get('action')

        with transaction.atomic():
            session = get_open_session(restaurant)

            if action == 'open':
                if session:
                    messages.error(request, 'Ya existe una caja abierta para este restaurante.')
                    return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

                opening_amount = parse_money(request.POST.get('opening_amount'))
                try:
                    session = CashRegisterSession.objects.create(
                        restaurant=restaurant,
                        opened_by=request.user,
                        opening_amount=opening_amount,
                    )
                except IntegrityError:
                    messages.error(request, 'Ya existe una caja abierta para este restaurante.')
                    return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

                CashMovement.objects.create(
                    restaurant=restaurant,
                    session=session,
                    user=request.user,
                    movement_type=CashMovement.TYPE_OPENING,
                    payment_method=CashMovement.PAYMENT_CASH,
                    amount=opening_amount,
                    note='Apertura de caja',
                )
                messages.success(request, 'Caja abierta correctamente.')

            elif action == 'movement' and session:
                movement_type = request.POST.get('movement_type')
                amount = parse_money(request.POST.get('amount'))
                note = request.POST.get('note', '').strip()
                payment_method = request.POST.get('payment_method') or CashMovement.PAYMENT_CASH

                if movement_type not in [CashMovement.TYPE_INCOME, CashMovement.TYPE_EXPENSE]:
                    messages.error(request, 'Tipo de movimiento inválido.')
                    return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

                if amount <= 0:
                    messages.error(request, 'El monto debe ser mayor a cero.')
                    return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

                if not note:
                    messages.error(request, 'El motivo es obligatorio.')
                    return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

                CashMovement.objects.create(
                    restaurant=restaurant,
                    session=session,
                    user=request.user,
                    movement_type=movement_type,
                    payment_method=payment_method,
                    amount=amount,
                    note=note,
                )
                messages.success(request, 'Egreso registrado correctamente.' if movement_type == CashMovement.TYPE_EXPENSE else 'Ingreso registrado correctamente.')

            elif action == 'close' and session:
                totals = get_session_totals(session)
                counted_cash = parse_money(request.POST.get('counted_cash') or request.POST.get('actual_total'))

                session.expected_total = totals['expected_total']
                session.actual_total = counted_cash
                session.expected_cash = totals['expected_cash']
                session.counted_cash = counted_cash
                session.difference = counted_cash - totals['expected_cash']
                session.total_cash = totals['cash_total']
                session.total_card = totals['card_total']
                session.total_qr = totals['digital_total']
                session.total_sales = totals['sales_total']
                session.total_tips = totals['tips_total']
                session.total_tax = totals['tax_total']
                session.total_expenses = totals['expense_total']
                session.notes = request.POST.get('notes', '').strip()
                session.closed_by = request.user
                session.closed_at = timezone.now()
                session.status = CashRegisterSession.STATUS_CLOSED
                session.save(update_fields=[
                    'expected_total',
                    'actual_total',
                    'expected_cash',
                    'counted_cash',
                    'difference',
                    'total_cash',
                    'total_card',
                    'total_qr',
                    'total_sales',
                    'total_tips',
                    'total_tax',
                    'total_expenses',
                    'notes',
                    'closed_by',
                    'closed_at',
                    'status',
                ])

                CashMovement.objects.create(
                    restaurant=restaurant,
                    session=session,
                    user=request.user,
                    movement_type=CashMovement.TYPE_CLOSING,
                    payment_method=CashMovement.PAYMENT_CASH,
                    amount=counted_cash,
                    note=session.notes or 'Cierre de caja',
                )
                messages.success(request, 'Caja cerrada correctamente.')

            else:
                messages.error(request, 'No hay caja abierta para esta operación.')

        return _dashboard_redirect(restaurant if user_is_superadmin(request.user) and not getattr(request, 'restaurant', None) else None)

    session = get_open_session(restaurant)
    last_closed_session = (
        CashRegisterSession.objects
        .filter(restaurant=restaurant, status=CashRegisterSession.STATUS_CLOSED)
        .select_related('opened_by', 'closed_by')
        .order_by('-closed_at')
        .first()
    )
    selected_session = session or last_closed_session
    totals, all_movements = _session_context(selected_session) if selected_session else (get_session_totals(None), CashMovement.objects.none())
    movements = all_movements[:20]

    context = {
        'restaurant': restaurant,
        'restaurants': restaurants,
        'is_global_cash_view': user_is_superadmin(request.user) and not getattr(request, 'restaurant', None),
        'session': session,
        'selected_session': selected_session,
        'last_closed_session': last_closed_session,
        'cash_total': totals['cash_total'],
        'card_total': totals['card_total'],
        'digital_total': totals['digital_total'],
        'qr_total': totals['qr_total'],
        'non_cash_total': totals['card_total'] + totals['digital_total'],
        'expected_total': totals['expected_total'],
        'expected_cash': totals['expected_cash'],
        'income_total': totals['income_total'],
        'income_cash': totals['income_cash'],
        'expense_total': totals['expense_total'],
        'expense_cash': totals['expense_cash'],
        'sales_total': totals['sales_total'],
        'tax_total': totals['tax_total'],
        'tips_total': totals['tips_total'],
        'movements': movements,
        'movement_types': [
            (CashMovement.TYPE_INCOME, 'Ingreso'),
            (CashMovement.TYPE_EXPENSE, 'Egreso'),
        ],
        'payment_methods': CashMovement.PAYMENT_METHOD_CHOICES,
    }

    return render(request, 'cash_register/index.html', context)


@login_required
@role_required(SUPERADMIN, ADMIN, CAJERO)
def export_cash_register_excel(request, session_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    session = _get_session_for_export(request, session_id)
    totals, movements = _session_context(session)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cierre caja'
    header_fill = PatternFill('solid', fgColor='00D4FF')
    header_font = Font(color='07111F', bold=True)

    summary_rows = [
        ('Restaurante', session.restaurant.name),
        ('Estado', session.get_status_display()),
        ('Apertura', timezone.localtime(session.opened_at).strftime('%Y-%m-%d %H:%M')),
        ('Cierre', timezone.localtime(session.closed_at).strftime('%Y-%m-%d %H:%M') if session.closed_at else '-'),
        ('Monto inicial', float(session.opening_amount)),
        ('Total efectivo', float(totals['cash_total'])),
        ('Total tarjeta', float(totals['card_total'])),
        ('Total QR', float(totals['digital_total'])),
        ('Efectivo esperado', float(session.expected_cash if session.status == CashRegisterSession.STATUS_CLOSED else totals['expected_cash'])),
        ('Efectivo contado', float(session.counted_cash if session.status == CashRegisterSession.STATUS_CLOSED else session.actual_total)),
        ('Diferencia', float(session.difference)),
        ('Ventas totales', float(totals['sales_total'])),
        ('IVA', float(totals['tax_total'])),
        ('Propinas', float(totals['tips_total'])),
        ('Egresos', float(totals['expense_total'])),
        ('Notas', session.notes or '-'),
    ]
    for row in summary_rows:
        ws.append(row)
    ws.append([])

    headers = ['Hora', 'Tipo', 'Metodo', 'Monto', 'Orden', 'Usuario', 'Motivo']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    for movement in movements:
        ws.append([
            timezone.localtime(movement.created_at).strftime('%Y-%m-%d %H:%M'),
            movement.get_movement_type_display(),
            movement.get_payment_method_display(),
            float(movement.amount),
            f'#{movement.order_id}' if movement.order_id else '-',
            movement.user.username,
            movement.note,
        ])

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'mesaflow-cierre-caja-{session.id}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
@role_required(SUPERADMIN, ADMIN, CAJERO)
def export_cash_register_pdf(request, session_id):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    session = _get_session_for_export(request, session_id)
    totals, movements = _session_context(session)
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('MesaFlow Cierre de Caja', styles['Title']),
        Paragraph(session.restaurant.name, styles['Heading2']),
        Paragraph(f'Apertura: {timezone.localtime(session.opened_at).strftime("%Y-%m-%d %H:%M")}', styles['Normal']),
        Paragraph(f'Cierre: {timezone.localtime(session.closed_at).strftime("%Y-%m-%d %H:%M") if session.closed_at else "-"}', styles['Normal']),
        Spacer(1, 16),
    ]

    summary_data = [
        ['Inicial', 'Efectivo', 'Tarjeta', 'QR', 'Esperado', 'Diferencia'],
        [
            f'${session.opening_amount:,.0f}',
            f'${totals["cash_total"]:,.0f}',
            f'${totals["card_total"]:,.0f}',
            f'${totals["digital_total"]:,.0f}',
            f'${(session.expected_cash if session.status == CashRegisterSession.STATUS_CLOSED else totals["expected_cash"]):,.0f}',
            f'${totals["sales_total"]:,.0f}',
            f'${session.difference:,.0f}',
        ],
    ]
    summary_table = Table(summary_data, colWidths=[72, 78, 78, 78, 78, 70])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00D4FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#07111F')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), .4, colors.HexColor('#B9C7D6')),
        ('PADDING', (0, 0), (-1, -1), 7),
    ]))
    story.extend([summary_table, Spacer(1, 18), Paragraph('Movimientos', styles['Heading2'])])

    movement_rows = [['Hora', 'Tipo', 'Metodo', 'Monto', 'Orden', 'Usuario']]
    for movement in movements[:35]:
        movement_rows.append([
            timezone.localtime(movement.created_at).strftime('%H:%M'),
            movement.get_movement_type_display(),
            movement.get_payment_method_display(),
            f'${movement.amount:,.0f}',
            f'#{movement.order_id}' if movement.order_id else '-',
            movement.user.username,
        ])
    if len(movement_rows) == 1:
        movement_rows.append(['Sin datos', '', '', '', '', ''])

    movements_table = Table(movement_rows, colWidths=[52, 82, 70, 78, 58, 95], repeatRows=1)
    movements_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#081028')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), .35, colors.HexColor('#B9C7D6')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(movements_table)
    doc.build(story)
    output.seek(0)

    filename = f'mesaflow-cierre-caja-{session.id}.pdf'
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
